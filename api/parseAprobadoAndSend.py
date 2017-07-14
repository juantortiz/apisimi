import csv
import sys
import os
import smtplib
import mimetypes
import mysql.connector
import ConfigParser
import openpyxl
from email.mime.multipart import MIMEMultipart
from email import encoders
from email.message import Message
from email.mime.audio import MIMEAudio
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.text import MIMEText
from dateutil.parser import parse as date_parse
from datetime import datetime, timedelta,date


attachs = []
dbuser = dbpass = db = emailfrom = emailto = username = password = smtp = dirtosearch =  ""

def init_config(configFile,emailfile):
    config = ConfigParser.ConfigParser()
    config.read(configFile)
    global emailfrom
    global emailto
    global username
    global password
    global smtp
    global dirtosearch
    global dbuser
    global dbpass
    global db
    global dbhost
    emailfrom = config.get('SMTP','emailfrom')
    username = config.get('SMTP','username')
    password = config.get('SMTP','password')
    smtp = config.get('SMTP','smtp')
    dirtosearch = config.get('GLOBAL','dirtosearch')
    dbuser = config.get('DB','dbusername')
    dbpass = config.get('DB','dbpassword')
    db = config.get('DB','db')
    dbhost = config.get('DB','dbhost')
    emailto = get_emails(emailfile)

def get_emails(filename):
    reader = csv.reader(open(filename))
    emails = []
    for row in reader:
        emails.append(str(row[0]))
    return emails

def get_fechas(fechadesde,fechahasta):
    fechas = []
    desde = datetime.strptime(fechadesde, '%Y-%m-%d_%H:%M:%S').date()
    hasta = datetime.strptime(fechahasta, '%Y-%m-%d_%H:%M:%S').date()
    delta = hasta - desde         # timedelta
    for i in range(delta.days + 1):
        aux = desde+timedelta(days=i)
        fechas.append(aux.strftime("%Y-%m-%d_00:00:00"))
    return fechas


def attach_files(msg):
    for i in attachs:
        attach_file(msg,i)


def attach_file(msg,filetoattach):
    ctype, encoding = mimetypes.guess_type(filetoattach)
    if ctype is None or encoding is not None:
        ctype = "application/octet-stream"

    maintype, subtype = ctype.split("/", 1)
    if maintype == "text":
        fp = open(filetoattach)
        # Note: we should handle calculating the charset
        attachment = MIMEText(fp.read(), _subtype=subtype)
        fp.close()
    else:
        fp = open(filetoattach, "rb")
        attachment = MIMEBase(maintype, subtype)
        attachment.set_payload(fp.read())
        fp.close()
        encoders.encode_base64(attachment)
    newfilename = filetoattach[filetoattach.rfind('/')+1:]
    attachment.add_header("Content-Disposition", "attachment", filename=newfilename)
    msg.attach(attachment)


def send_mails(msg):
    server = smtplib.SMTP(smtp)
    server.starttls()
    server.login(username,password)
    server.sendmail(emailfrom, emailto, msg.as_string())
    server.quit()

def create_mails(dia,estado):
    msg = MIMEMultipart()
    msg["From"] = emailfrom
    msg["To"] = ', '.join(emailto)
    subject = "Resumen de simis en el estado "+estado+" al dia " if (len(attachs) > 0) else "No hay simis aprobadas en el dia "
    msg["Subject"] = subject + dia[:dia.find('_')]
    msg.preamble = "Ver adjuntos"
    attach_files(msg)
    return msg

def get_archivos(fechas,dirToSearch):
    archivos = []
    for fecha in fechas:
        fecha = fecha[:fecha.find('_')]
        for i in os.listdir(dirToSearch):
            if os.path.isfile(os.path.join(dirToSearch,i)) and fecha in i:
                if 'A1dest' not in i:
                    archivos.append((dirToSearch+'/'+i,dirToSearch+'/salidasUsuario/'+i))
    return archivos

def a_letra(estado):
    if estado == 'APROBADA':
        return 'A'
    elif estado == 'CAUTELAR':
        return 'C'
    else:
        return ''

def escribir_archivos(archivos,estado):
    for archivo in archivos:
        escribir_archivo(archivo,estado)

def get_cuit(simi):
    cnx = mysql.connector.connect(user=dbuser,password=dbpass, database=db, host = dbhost)
    cursor = cnx.cursor(buffered=True)
    query = ("select vil2.value from ProcessInstanceLog as pil join VariableInstanceLog as vil on vil.processinstanceid = pil.processinstanceid and vil.variableId = 'estado_simi' join VariableInstanceLog as vil1 on vil1.processinstanceid = pil.processinstanceid and vil1.variableId='djai_id_simi'  join VariableInstanceLog as vil2 on vil2.processinstanceid = pil.processinstanceid and vil2.variableId='djai_cuit_imp' where vil1.value = '"+simi+"' and vil.value='CAUTELAR';")
    cursor.execute(query,simi)
    leg = cursor.fetchone()
    cursor.close()
    cnx.close()
    return leg

def esta_realmente_aprobada(simi,estado):
    cnx = mysql.connector.connect(user=dbuser,password=dbpass, database=db, host = dbhost)
    cursor = cnx.cursor(buffered=True)
    query = ("select DISTINCT pil.id FROM ProcessInstanceLog as pil join VariableInstanceLog as vil on vil.processinstanceid = pil.processinstanceid and vil.variableId = 'estado_simi'"
                    "join VariableInstanceLog as vil1 on vil1.processinstanceid = pil.processinstanceid and vil1.variableId='djai_id_simi'"
                    "where vil1.value = '"+simi+"' and vil.value='"+estado+"';")

    cursor.execute(query,simi)
    leg_no = cursor.fetchall()
    cursor.close()
    cnx.close()
    if len(leg_no) > 0:
        return True
    else:
        return False

def get_archivo_salida(archivo,estado,key):
    archivo = archivo[:-3]
    raiz = archivo[archivo.rfind('/')+1:]
    nombre = raiz + key +'.'+estado+ '.xlsx'
    path = archivo+key+'.'+estado+'.xlsx'
    return nombre,path

def escribir_simis(simis,estado,archivo):
    for key in simis.keys():
        wb = openpyxl.Workbook()
        ws = wb.get_sheet_by_name('Sheet')
        ws.title = 'Hoja1'
        archivoSalida,path = get_archivo_salida(archivo,estado,key)
        aux = simis[key]
        for idx,simi in enumerate(aux):
            ws.cell(row=idx+1,column=1).value = simi['simi']
            if simi.has_key('cuit'):
                ws.cell(row=idx+1,column=2).value = simi['cuit']
        attachs.append(path)
        wb.save(path)

def escribir_archivo(archivo,estado):
    simis = get_simis_from_file(estado,archivo[0])
    escribir_simis(simis,estado,archivo[1])

def get_simis_from_file(estado,archivo):
    letraEstado = a_letra(estado)
    reader = csv.reader(open(archivo),delimiter=',')
    simis = {}
    for row in reader:
        cuit = None
        if row[1] == letraEstado:
            if esta_realmente_aprobada(row[0],estado):
                aux = {'simi':row[0]}
                if letraEstado == 'C':
                    cuit = get_cuit(row[0])
                    aux['cuit'] = cuit[0]
                nombre = row[3]
                if not simis.has_key(nombre):
                    simis[nombre] = []
                simis[nombre].append(aux)
    return simis

if __name__ == '__main__':
    if len(sys.argv) != 6 :
        print("Argumentos incorrectos")
        sys.exit()
    else:
        estado = sys.argv[4]
        if estado != 'APROBADA' and estado != 'CAUTELAR':
            print("Los posibles estados son APROBADA|CAUTELAR")
            sys.exit()
        init_config(sys.argv[1],sys.argv[5])
        fechadesde = sys.argv[2]
        fechahasta = sys.argv[3]
        fechas = get_fechas(fechadesde,fechahasta)
        archivos = get_archivos(fechas,dirtosearch)
        escribir_archivos(archivos,estado)
        send_mails(create_mails(fechas[-1],estado))


